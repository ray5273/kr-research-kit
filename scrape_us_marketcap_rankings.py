"""
미국 상장기업 연도별 시가총액 상위 1,000개 수집 스크립트
=========================================================
데이터 출처 : Atlantis Data Solutions (https://atlantisdatasolutions.com/rankings)
원천 데이터 : SEC 공시(XBRL) 기반, 각 연도 연말(calendar year-end) 시가총액
커버 범위  : 2010-2025 각 1,000개 + Live(실시간 스냅샷) + 2006(일부만 존재)

실행 방법:
    pip install requests pandas openpyxl
    python scrape_us_marketcap_rankings.py

출력 파일:
    us_marketcap_rankings.csv          (long-format 통합 CSV)
    us_marketcap_top1000_by_year.xlsx  (연도별 시트 엑셀)

주의 사항:
    - 2009년 데이터는 소스에 없음 (SEC XBRL 의무공시가 2010년 전후 도입)
    - 하위 순위권에는 주식수 자동 파싱 오류로 보이는 이상치가 일부 존재
    - 페이지 구조(테이블 형태)가 바뀌면 파싱 로직 수정 필요
"""

import re
import csv
import requests
import pandas as pd
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

URL = "https://atlantisdatasolutions.com/rankings"
CSV_OUT = "us_marketcap_rankings.csv"
XLSX_OUT = "us_marketcap_top1000_by_year.xlsx"


# ---------------------------------------------------------------
# 1단계. 페이지 다운로드
# ---------------------------------------------------------------
def fetch_html(url: str = URL) -> str:
    """순위 페이지 전체 HTML을 받아온다 (약 2.5MB, 서버사이드 렌더링이라 JS 불필요)."""
    resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=60)
    resp.raise_for_status()
    return resp.text


# ---------------------------------------------------------------
# 2단계. HTML 테이블 파싱
# ---------------------------------------------------------------
# 테이블 구조:
#   <thead>에 연도 헤더가 4개씩 반복됨 (Rank / Company / Market Cap / % of Total)
#   <tbody>의 각 <tr>이 "순위 1줄" = 모든 연도의 해당 순위 기업이 가로로 나열
#   회사 셀 형태: <a href="/aapl/summary">AAPL</a><br><small>Apple</small>
def parse_rankings(html: str) -> list[tuple]:
    """HTML에서 (연도, 순위, 티커, 회사명, 시총B, 비중%) 레코드 리스트를 뽑는다."""
    head = re.search(r"<thead>(.*?)</thead>", html, re.S).group(1)
    ths = re.findall(r"<th>([^<]*)", head)
    years = [y.strip() for y in ths[::4] if y.strip()]  # 4칸마다 연도 1개

    body = re.search(r"<tbody>(.*?)</tbody>", html, re.S).group(1)
    rows = re.findall(r"<tr>(.*?)</tr>", body, re.S)

    cell_re = re.compile(r"<td>(.*?)</td>", re.S)
    comp_re = re.compile(r">([^<]+)</a>.*?<small>([^<]*)</small>", re.S)

    records = []
    for row in rows:
        cells = cell_re.findall(row)
        if len(cells) != len(years) * 4:
            continue  # 구조가 깨진 행은 건너뜀
        for i, year in enumerate(years):
            rank_raw, comp_raw, mcap_raw, pct_raw = cells[i * 4 : i * 4 + 4]
            rank_raw = rank_raw.strip()
            if not rank_raw:  # 해당 연도에 데이터가 없는 칸 (예: 2006)
                continue
            m = comp_re.search(comp_raw)
            if not m:
                continue
            ticker, name = m.group(1).strip(), m.group(2).strip()
            try:
                rank = int(float(rank_raw))                              # "1.0" -> 1
                mcap = float(mcap_raw.strip().replace(",", "").rstrip("B"))  # "5,142B" -> 5142.0
            except ValueError:
                continue
            pct = pct_raw.strip().rstrip("%")
            records.append((year, rank, ticker, name, mcap, pct))
    return records


# ---------------------------------------------------------------
# 3단계. CSV 저장 (long-format)
# ---------------------------------------------------------------
def save_csv(records: list[tuple], path: str = CSV_OUT) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["year", "rank", "ticker", "company",
                    "market_cap_billion_usd", "pct_of_total_market"])
        for rec in sorted(records, key=lambda r: (r[0], r[1])):
            w.writerow(rec)


# ---------------------------------------------------------------
# 4단계. 엑셀 저장 (연도별 시트)
# ---------------------------------------------------------------
def save_xlsx(records: list[tuple], path: str = XLSX_OUT) -> None:
    df = pd.DataFrame(records, columns=["year", "rank", "ticker", "company",
                                        "market_cap_billion_usd", "pct_of_total_market"])
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")

    year_order = [str(y) for y in range(2025, 2009, -1)] + ["Live", "2006"]
    for year in year_order:
        sub = df[df["year"].astype(str) == year].sort_values("rank")
        if sub.empty:
            continue
        ws = wb.create_sheet(year)
        ws.append(["rank", "ticker", "company",
                   "market_cap_billion_usd", "pct_of_total_market"])
        for _, r in sub.iterrows():
            ws.append([int(r["rank"]), r["ticker"], r["company"],
                       float(r["market_cap_billion_usd"]), r["pct_of_total_market"]])
        for c in ws[1]:
            c.fill, c.font = hdr_fill, hdr_font
            c.alignment = Alignment(horizontal="center")
        for col, width in zip("ABCDE", [8, 10, 38, 22, 18]):
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font = Font(name="Arial", size=10)
            row[3].number_format = "#,##0.0"
        ws.freeze_panes = "A2"
    wb.save(path)


# ---------------------------------------------------------------
# 실행부
# ---------------------------------------------------------------
if __name__ == "__main__":
    html = fetch_html()
    records = parse_rankings(html)

    counts = Counter(r[0] for r in records)
    print(f"총 레코드: {len(records)}")
    for year, n in sorted(counts.items()):
        print(f"  {year}: {n}개")

    save_csv(records)
    save_xlsx(records)
    print(f"저장 완료: {CSV_OUT}, {XLSX_OUT}")
