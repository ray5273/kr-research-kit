# -*- coding: utf-8 -*-
"""
한국 주식시장(KOSPI+KOSDAQ) 연도별 시가총액 상위 N개 종목 추출 스크립트
====================================================================

데이터 출처
- 원천: 한국거래소(KRX) 정보데이터시스템 (data.krx.co.kr)
- 경유: FinanceData/marcap 데이터셋 (https://github.com/FinanceData/marcap)
  KRX 일별 시세를 1995년부터 현재까지 수집해 연도별 parquet/csv로 배포하는 오픈소스

필요 패키지
    pip install pandas pyarrow openpyxl

사용법
    python extract_top300_marcap.py
    (아래 [설정] 부분에서 연도 범위, 상위 개수 등을 수정)
"""

import os
import re
import subprocess
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# [설정]
# ─────────────────────────────────────────
START_YEAR = 2015          # 2016년 운용 유니버스용 전년말 스냅샷 포함
END_YEAR = 2025            # 종료 연도
TOP_N = 300                # 상위 몇 개 종목을 뽑을지
INCLUDE_MARKETS = ['KOSPI', 'KOSDAQ', 'KOSDAQ GLOBAL']  # KONEX 제외
# Annual strategy universes are common shares only.  Keep this default in the
# extractor as well as in the JS loader so a workbook cannot silently re-add
# ETFs, ETNs, REITs, SPACs, or preferred shares.
EXCLUDE_PREFERRED = True
MARCAP_DIR = 'marcap'      # marcap 저장소를 받을 폴더
OUTPUT_FILE = f'한국_시가총액_상위{TOP_N}_{START_YEAR}-{END_YEAR}.xlsx'

# ─────────────────────────────────────────
# 1) marcap 데이터셋 준비 (없으면 git clone)
# ─────────────────────────────────────────
def prepare_marcap():
    if not os.path.isdir(os.path.join(MARCAP_DIR, 'data')):
        print('marcap 데이터셋 다운로드 중...')
        subprocess.run(
            ['git', 'clone', '--depth', '1',
             'https://github.com/FinanceData/marcap.git', MARCAP_DIR],
            check=True
        )
    else:
        print('marcap 데이터셋이 이미 존재합니다. (최신화하려면 폴더 삭제 후 재실행)')


# ─────────────────────────────────────────
# 2) 연도별 마지막 거래일 기준 상위 N개 추출
# ─────────────────────────────────────────
def is_common_share(row) -> bool:
    """Return True only for an ordinary KOSPI/KOSDAQ operating share.

    A code-ending heuristic alone is not sufficient: some common shares do
    not end in zero and fund-like products can.  Names are deliberately
    filtered conservatively; the normalized ledger retains the raw count so
    exclusions remain auditable.
    """
    name = str(row.get('Name', ''))
    market = str(row.get('Market', ''))
    if market not in INCLUDE_MARKETS:
        return False
    excluded = ('ETF', 'ETN', '스팩', 'SPAC', '우선', '우B',
                '레버리지', '인버스', '커버드콜')
    # 메리츠금융지주처럼 이름 중간에 “리츠”가 들어간 보통주를 제외하지
    # 않는다. 실제 상장 리츠는 통상 이름이 리츠로 끝난다.
    is_reit = name.strip().endswith('리츠') or 'REIT' in name.upper()
    is_preferred = bool(re.search(r'우선|[1-3]?우(?:B|C)?$', name))
    return not any(token.lower() in name.lower() for token in excluded) and not is_reit and not is_preferred


def extract_year(year: int) -> tuple:
    path = os.path.join(MARCAP_DIR, 'data', f'marcap-{year}.parquet')
    df = pd.read_parquet(path)

    # 시장 필터 (KONEX 제외)
    df = df[df['Market'].isin(INCLUDE_MARKETS)]

    # 해당 연도 마지막 거래일 스냅샷
    last_date = df['Date'].max()
    snap = df[df['Date'] == last_date].copy()

    # 보통주/운영회사 주식만 남긴 뒤 순위를 다시 부여한다.
    if EXCLUDE_PREFERRED:
        snap = snap[snap.apply(is_common_share, axis=1)]

    # 시가총액 내림차순 상위 N
    snap = snap.sort_values('Marcap', ascending=False).head(TOP_N).reset_index(drop=True)
    snap['Market'] = snap['Market'].replace({'KOSDAQ GLOBAL': 'KOSDAQ'})
    return last_date, snap


# ─────────────────────────────────────────
# 3) 엑셀 파일 생성
# ─────────────────────────────────────────
def build_excel(results: dict):
    wb = Workbook()
    thin = Border(bottom=Side(style='thin', color='D9D9D9'))
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    base_font = Font(name='Arial', size=10)

    # ── 개요 시트
    ws = wb.active
    ws.title = '개요'
    ws['A1'] = (f'한국 주식시장(KOSPI+KOSDAQ) 연도별 시가총액 상위 {TOP_N} 종목 '
                f'({START_YEAR}-{END_YEAR})')
    ws['A1'].font = Font(name='Arial', bold=True, size=13)
    ws['A3'] = ('기준: 각 연도 마지막 거래일 종가 기준 / KONEX 제외'
                + (' / ETF·ETN·리츠·스팩·우선주 제외' if EXCLUDE_PREFERRED else ' / 전 종목 포함'))
    ws['A4'] = ('출처: 한국거래소(KRX) 정보데이터시스템 일별 시세 '
                '(FinanceData marcap 데이터셋 경유, github.com/FinanceData/marcap)')
    ws['A5'] = '단위: 종가(원), 시가총액(억원), 상장주식수(주)'
    for r in (3, 4, 5):
        ws.cell(row=r, column=1).font = Font(name='Arial', size=10,
                                             italic=True, color='595959')

    headers = ['연도', '기준일(마지막 거래일)', '시총 1위 종목', '1위 시가총액(조원)']
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=7, column=j, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
    for i, (year, (d, snap)) in enumerate(sorted(results.items())):
        r = 8 + i
        ws.cell(row=r, column=1, value=year).font = base_font
        ws.cell(row=r, column=2, value=d.strftime('%Y-%m-%d')).font = base_font
        ws.cell(row=r, column=3, value=snap.iloc[0]['Name']).font = base_font
        c = ws.cell(row=r, column=4, value=round(snap.iloc[0]['Marcap'] / 1e12, 1))
        c.font = base_font
        c.number_format = '#,##0.0'
    for col, w in zip('ABCD', [10, 22, 18, 18]):
        ws.column_dimensions[col].width = w

    # ── 연도별 시트
    cols = ['순위', '종목코드', '종목명', '시장', '종가(원)', '시가총액(억원)', '상장주식수']
    for year in sorted(results):
        d, snap = results[year]
        ws = wb.create_sheet(str(year))
        ws.cell(row=1, column=1,
                value=f'{year}년 시가총액 상위 {TOP_N} '
                      f'(기준일: {d.strftime("%Y-%m-%d")}, KOSPI+KOSDAQ)'
                ).font = Font(name='Arial', bold=True, size=11)
        for j, h in enumerate(cols, 1):
            c = ws.cell(row=3, column=j, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center')
        for i, row in snap.iterrows():
            r = 4 + i
            vals = [i + 1, row['Code'], row['Name'], row['Market'],
                    int(row['Close']), round(row['Marcap'] / 1e8, 0),
                    int(row['Stocks'])]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = base_font
                c.border = thin
                if j in (1, 2):
                    c.alignment = Alignment(horizontal='center')
                if j in (5, 6, 7):
                    c.number_format = '#,##0'
        for col_idx, w in zip(range(1, 8), [7, 11, 22, 10, 12, 15, 15]):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.freeze_panes = 'A4'

    wb.save(OUTPUT_FILE)
    print(f'저장 완료: {OUTPUT_FILE}')


# ─────────────────────────────────────────
# 실행
# ─────────────────────────────────────────
if __name__ == '__main__':
    prepare_marcap()
    results = {}
    for y in range(START_YEAR, END_YEAR + 1):
        d, snap = extract_year(y)
        results[y] = (d, snap)
        top1 = snap.iloc[0]
        print(f'{y} | 기준일 {d.date()} | {len(snap)}종목 | '
              f'1위: {top1["Name"]} {top1["Marcap"] / 1e12:.1f}조')
    build_excel(results)
